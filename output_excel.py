from openpyxl import Workbook, load_workbook
from CONFIG import *
from openpyxl.styles import PatternFill, Side, Border
import datetime
import timeit


class ExcelOutData:
    """ Create new sheet and print out all useful raw data to the user"""
    def __init__(self, scheme, time_start):
        self.sch = scheme
        self.led = self.sch.led
        self.time_start = time_start

        self.wb = load_workbook(sh_file)
        self.ws = self.wb.create_sheet(sh_data, 1)
        self.ws3 = self.wb.create_sheet(sh_hours, 1)

        """ Metod exe """
        self.add_dates(self.ws)
        self.add_block_main()
        self.create_fail_log()
        self.add_block_hours()
        self.hour_diff()
        self.add_block_underh()
        self.add_block_typescounter_hl()
        self.add_block_typescounter_ul()
        self.add_loop_count()
        self.add_file_name()

        self.add_dates(self.ws3)
        self.full_hours_block()
        self.col_width()
        self.color_scheme()
        self.ws3.append([''])
        self.ws3.append(['Använd detta blad för att klistra in tiderna i lägerschemat'])
        self.ws3.append(['Fungerar bäst in i Word och inte Docs'])
        self.save()
        print("\nKlart! Schemat finns i excelfilen:", sh_file)

    def get_hour_text(self, type, led):
        if led.isHl:
            if type == 'A':
                return A_hl_txt
            elif type == 'B':
                return B_hl_txt
            elif type == 'C':
                return C_hl_txt
            elif type == 'D':
                return D_hl_txt
            elif type == 'L':
                return 'Ledig'
        elif led.isHl is False:
            if type == 'A':
                return A_ul_txt
            elif type == 'B':
                return B_ul_txt
            elif type == 'C':
                return C_ul_txt
            elif type == 'D':
                return D_ul_txt
            elif type == 'L':
                return 'Ledig'

    def full_hours_block(self):
        for led in self.led:
            line = []
            for type in led.total:
                line.append(' ')
                line.append(led.name)
                line.append(self.get_hour_text(type, led))
            self.ws3.append(line)

    def color_scheme(self):
        """ loop trough all cells """
        for row in self.ws3.rows:
            for cell in row:
                self.set_color(cell)

    def set_color(self, cell):
        """ Add color and border style to cell """
        if cell.value in [A_hl_txt, A_ul_txt]:
            cell.font = Atxt
            cell.fill = BGcol

            '''
        elif cell.value in [B_hl_txt, B_ul_txt]:
            cell.fill = Bcol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        elif cell.value in [C_hl_txt, C_ul_txt]:
            cell.fill = Ccol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)'''
        elif cell.value in [D_hl_txt, D_ul_txt]:
            cell.font = Dtxt
            cell.fill = BGcol

        elif cell.value == 'Ledig':
            cell.font = Ltxt
            cell.fill = BGcol

        else:
            cell.fill = BGcol

        cell.border = Border(top=bd_gray, left=bd_gray, right=bd_gray, bottom=bd_gray)

    def hlorulword(self, led):
        """ dumb ass name"""
        if led.isHl is True:
            return "hl"
        else:
            return "ul"

    def add_block_main(self):
        """ Scheme block with all shifts, names and tot hours"""
        for led in self.led:
            line = led.total.copy()
            line.insert(0, led.name)
            line.insert(0, self.hlorulword(led))
            line.append('')
            line.append('Timmar:')
            line.append(led.hours_tot)
            self.add_line(line)
        self.add_line([])

    def add_block_hours(self):
        """ Hours block with hours per shift """
        self.add_line(['Antalet timmar per pass'])
        for led in self.led:
            line = led.hours.copy()
            line.insert(0, led.name)
            line.insert(0, self.hlorulword(led))
            line.append('')
            line.append('Timmar:')
            line.append(led.hours_tot)
            self.add_line(line)
        self.add_line([])

    def add_block_typescounter_hl(self):
        """ Type counter block (hl) """
        for type in ['A', 'B', 'C', 'D', 'L']:
            if type == 'A':
                line = self.sch.hlA_count.copy()
            elif type == 'B':
                line = self.sch.hlB_count.copy()
            elif type == 'C':
                line = self.sch.hlC_count.copy()
            elif type == 'D':
                line = self.sch.hlD_count.copy()
            elif type == 'L':
                line = self.sch.hlL_count.copy()

            line.insert(0, f"Hl {type}")
            line.insert(0, 'Antal:')

            self.add_line(line)
        self.add_line([])

    def add_block_typescounter_ul(self):
        """ Type counter block (ul) """
        for type in ['A', 'B', 'C', 'D', 'L']:
            if type == 'A':
                line = self.sch.A_count.copy()
            elif type == 'B':
                line = self.sch.B_count.copy()
            elif type == 'C':
                line = self.sch.C_count.copy()
            elif type == 'D':
                line = self.sch.D_count.copy()
            elif type == 'L':
                line = self.sch.L_count.copy()

            line.insert(0, f"Ul {type}")
            line.insert(0, 'Antal:')

            self.add_line(line)
        self.add_line([])

    def add_dates(self, ws):
        """ Header for blocks """
        self.add_line([])

        line = []
        line.append('Lägerdag:')
        line.append('')
        start = datetime.date(2022, 7, 21)
        for it in range(len(self.led[0].total)):
            dt = datetime.timedelta(it)
            msg = "{:%a (%d/%m)}"
            if date_info is True:
                line.append(msg.format(start + dt))  # check config to change from "Lägerdag x" to dates
            else:
                line.append(f"Lägerdag {it + 1}")
            if ws == self.ws3:
                line.append('')
                line.append('')

        if ws == self.ws:
            self.add_line(line)
        elif ws == self.ws3:
            self.ws3.append(line)

    def col_width(self):  # dont call this
        """ Adjusts the coloumn width. dont call this"""
        for col in self.ws3.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)
            self.ws3.column_dimensions[column].width = adjusted_width


    def add_block_underh(self):
        """ Under houred block (bools) """
        self.add_line(['Har ledaren arbetet för få timmar? '])
        self.add_line(['Falsk betyder att ledaren har fler total timmar än snittet, Sant så har ledaren för få timmar'])
        for led in self.led:
            line = led.isUnderh.copy()
            line.insert(0, led.name)
            line.insert(0, self.hlorulword(led))
            line.append('')
            line.append('Timmar:')
            line.append(led.hours_tot)
            self.add_line(line)
        self.add_line([])

    def hour_diff(self):
        """ print the hour difference for most houred worked to least"""
        tot_hour_list_hl = []
        tot_hour_list_ul = []
        for led in self.led:
            if led.isHl:
                tot_hour_list_hl.append(led.hours_tot)
            else:
                tot_hour_list_ul.append(led.hours_tot)

        ul_diff = max(tot_hour_list_ul) - min(tot_hour_list_ul)
        hl_diff = max(tot_hour_list_hl) - min(tot_hour_list_hl)
        self.add_line(["Skillnaden i totala antal timmar jobbade:"])
        self.add_line([f"Huvudledarna: {hl_diff} timmar i skillnad"])
        self.add_line([f"Ungdomsledarna: {ul_diff} timmar i skillnad"])
        self.add_line([])
        print("\nSkillnaden i totala antal timmar jobbade:")
        print(f"Huvudledarna: {hl_diff} timmar i skillnad")
        print(f"Ungdomsledarna: {ul_diff} timmar i skillnad")

    def add_loop_count(self):
        """ Number of loops and time the algo needed to create scheme"""
        time_stop = timeit.default_timer()
        self.add_line([f'Antalet loops som programmet försökte placera ut pass på: {self.sch.algo_loop_counter} '])
        self.add_line([f'Tid för att beräkna nytt schema: {round((time_stop - self.time_start), 3)} sek'])
        self.add_line([])

    def add_file_name(self):
        self.add_line([f'Excelfil som schemat baseras på: {cfile}'])
        self.add_line([])

    def create_fail_log(self):
        """ If placeralgo couldn't place shift, print a error msg"""
        self.add_line(['FELMEDDELANDE'])
        for msg in self.sch.fail_log:
            self.add_line(msg)
        self.add_line([])

    def add_line(self, line):
        self.ws.append(line)


    def save(self):
        self.wb.save(sh_file)


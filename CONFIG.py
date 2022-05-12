from openpyxl.styles import Font, PatternFill, Side
from openpyxl import styles


###########################################
#
#         KONFIGURATIONSFILEN
#
# I den här filen kan du ändra på vissa värden.
#
# Det du kan ändra på är detta:
# Arbetstimmarna, max antal pass, datum för lägret, namn på filer och blad
# Färger och styles, excelfil som används, hur många loops som tillåts
#
# Var noggrann med vad du ändrar, en felformatering kan leda till att programmet kraschar
# Allt med en # innan är endast kommentarer och påverkar inte programmet.
# Desto längre ner i dokumentet du kommer, desto mindre relevanta blir ändringarna
#
# Se till att spara filen efter du ändrat något
#
###########################################







#    ARBETSPASS UL

# Antalet timmar per pass
# Ändra dessa om du vill ändra tiderna för passen för ungdomsledarna
Ah = 9.5
Bh = 12.5
Ch = 7.5
Dh = 11

# Mellan vilka timmar
# OBS dessa är endast text för kosmetisk utskrift, påverkar inte programmet
A_ul_txt = '8:30-19:00'
B_ul_txt = '10:00-22:30'
C_ul_txt = '13:00-20:30'
D_ul_txt = '15:00-01:00'


#   ARBETSPASS HL


# Antalet timmar
# Ändra dessa om du vill ändra tiderna för passen för huvudledarna
Ahhl = 12
Bhhl = 12.5
Chhl = 7.5
Dhhl = 11

# Mellan vilka timmar
# OBS dessa är endast text för kosmetisk utskrift, påverkar inte programmet
A_hl_txt = '8:30-20:30'
B_hl_txt = '09:30-21:00'
C_hl_txt = '13:00-22:30'
D_hl_txt = '14:30-01:30'



# Max antal pass per dag
# Ändra dessa om du vill tillåta mer/mindre antal pass för UL
amax = 2
# bmax = 2   Går inte att ändra på här eftersom det beror på antalet ledare
# cmax = 2   Går endast att ändra i place_algo_methods.py genom att ändra på koden
dmax = 2

# Ändra dessa om du vill tillåta mer/mindre antal pass för HL
hlamax = 1
# hlbmax = 2  går ej att ändra
# hlcmax = 2  går ej att ändra
hldmax = 1




# Start datum för lägret
# Ändra dessa om du vill ha datum som rubriker istället för lägerdagar
year = 2022
month = 7
day = 18

# Sätt till True om du vill visa datum istället
date_info = False


# Excelfil som programmet utgår ifrån (config file)
cfile = 'TOMT_SCHEMA.xlsx'

# Namn på filer och excelblad
sh_data = 'Rå data'
sh_hours = 'Arbetstimmar'
sh_final = 'Schema'
sh_file = 'LÄGERSCHEMA.xlsx'



# Färger
# Ändra endast hex-värdet för att byta färg
beige = styles.Color(rgb='FFE598', tint=0.0)
green = styles.Color(rgb='C5E0B3', tint=0.0)
pink = styles.Color(rgb='FF8AD8', tint=0.0)
blue = styles.Color(rgb='B4C6E7', tint=0.0)
yellow = styles.Color(rgb='FFC000', tint=0.0)
bggray = styles.Color(rgb='f4f4f4', tint=0.2)
darkbeige = styles.Color(rgb='FFD24C', tint=0.0)
darkblue = styles.Color(rgb='7A9AD4', tint=0.0)
darkyellow = styles.Color(rgb='ce7e00', tint=0.0)




# Styles till excel, tjockleken på kanterna
bd = Side(border_style='thin', color="000000")
bd_thick = Side(border_style='thick', color="000000")
bd_gray = Side(border_style='thin', color="999999")



# Styles till excel, färger
Acol = PatternFill(fill_type="solid", fgColor=beige)
Bcol = PatternFill(fill_type="solid", fgColor=pink)
Ccol = PatternFill(fill_type="solid", fgColor=green)
Dcol = PatternFill(fill_type="solid", fgColor=blue)
Lcol = PatternFill(fill_type="solid", fgColor=yellow)
BGcol = PatternFill(fill_type="solid", fgColor=bggray)

Atxt = Font(bold=False, color=darkbeige)
Btxt = Font(bold=False, color=pink)
Ctxt = Font(bold=False, color=green)
Dtxt = Font(bold=False, color=darkblue)
Ltxt = Font(bold=False, color=darkyellow)
BGtxt = Font(bold=False, color=bggray)


# Max antal försök programmet får för lösa ett problem
# Behöver inte ändras på
max_fail_count = 60

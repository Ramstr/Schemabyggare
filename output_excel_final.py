from openpyxl import Workbook, load_workbook
from CONFIG import *
from openpyxl.styles import PatternFill, Side, Border
import datetime


class ExcelOutMain:
    """ Create a excel file and print a coloured scheme """
    def __init__(self, scheme):
        self.sch = scheme
        self.led = self.sch.led

        self.wb = Workbook()
        self.ws2 = self.wb.active
        self.ws2.title = sh_final

        """ Method exe """
        self.add_dates()
        self.add_block_main()
        self.color_scheme()
        self.col_width()
        self.save()

    def hlorulword(self, led):
        """ dumb ass name"""
        if led.isHl is True:
            return "hl"
        else:
            return "ul"

    def add_block_main(self):
        """ Scheme block with all shifts, names and tot hours"""
        for led in self.led:
            line = led.total.copy()
            line.insert(0, led.name)
            # line.insert(0, self.hlorulword(led))
            line.append('')
            line.append('Timmar:')
            line.append(led.hours_tot)
            self.add_line(line)
        self.add_line([])

    def add_line(self, line):
        self.ws2.append(line)

    def add_dates(self):
        """ Header for the scheme """
        line = []
        line.append('Lägerdag:')
        start = datetime.date(year, month, day)
        for it in range(len(self.led[0].total)):
            dt = datetime.timedelta(it)
            msg = "{:%a (%d/%m)}"
            if date_info is True:  # check config to change from "Lägerdag x" to dates
                line.append(msg.format(start + dt))
            else:
                line.append(f"Lägerdag {it + 1}")
        self.add_line(line)

    def col_width(self):
        """ Adjusts the coloumn width """
        for col in self.ws2.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)
            self.ws2.column_dimensions[column].width = adjusted_width

    def color_scheme(self):
        """ loop trough all cells """
        for row in self.ws2.rows:
            for cell in row:
                self.set_color(cell)

    def set_color(self, cell):
        """ Add color and border style to cell """
        if cell.value == 'A':
            cell.fill = Acol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        elif cell.value == 'B':
            cell.fill = Bcol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        elif cell.value == 'C':
            cell.fill = Ccol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        elif cell.value == 'D':
            cell.fill = Dcol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        elif cell.value == 'L':
            cell.fill = Lcol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        else:
            cell.fill = BGcol
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)

    def save(self):
        self.wb.save(sh_file)

from openpyxl import load_workbook
from ledare_object import *
from CONFIG import *


class ExcelIn:
    """
    Load excel and get data from user
    """
    def __init__(self):
        self.wb = load_workbook(filename=cfile)  # Loads excel file
        self.led = []  # list for all leader-objects
        self.ws = self.wb.active  # wb=workbook, ws=worksheet

        """ Method exe """
        self.load_excel()
        self.load_hours()
        self.calc_total_hours()

        """ Terminal Outputs """
        # self.output_hours()
        # self.output_aval()

    def load_excel(self):
        """
        Load in from excel
        """
        def chech_hl(y):
            """
            Determain if led is hl or ul
            :param y: cell-object
            :return: bool
            """
            if self.ws[f'A{y + 1}'].value == 'hl':
                return True
            else:
                return False

        for i in range(self.led_range()):
            self.led.append(Led2(self.ws[f'B{i + 1}'].value, chech_hl(i), i))  # Create led-object
            for cell in self.ws[f'{i + 1}']:
                if cell.value in ['A', 'B', 'C', 'D', 'Fa', 'Fb', 'Fc', 'Fd', '_', 'L']:
                    self.led[i].total.append(cell.value)
                else:
                    pass

    def led_range(self):
        return len(self.ws['A'])

    def led_index(self):
        """
        Calc max/min index of hl and ul
        """
        ul_count = 0
        hl_count = 0
        for i in self.led:
            if i.isHl is False:
                ul_count += 1
            elif i.isHl is True:
                hl_count += 1
        return hl_count, hl_count + ul_count - 1

    def load_hours(self):
        """ Load all led shifts """
        for row in range(len(self.led[0].total)):
            for col, led in enumerate(self.led):
                self.add(row, col, led.total[row])

    def add(self, col, row, type):
        """ Add the hours for all shifts """
        self.led[row].available.append(False)

        if type == "A" or type == "Fa":
            if self.led[row].isHl is False:
                self.led[row].hours.append(Ah)

            elif self.led[row].isHl:
                self.led[row].hours.append(Ahhl)

        elif type == "B" or type == "Fb":
            if self.led[row].isHl is False:
                self.led[row].hours.append(Bh)

            elif self.led[row].isHl:
                self.led[row].hours.append(Bhhl)

        elif type == "C" or type == "Fc":

            if self.led[row].isHl is False:
                self.led[row].hours.append(Ch)

            elif self.led[row].isHl:
                self.led[row].hours.append(Chhl)

        elif type == "D" or type == "Fd":

            if self.led[row].isHl is False:
                self.led[row].hours.append(Dh)

            elif self.led[row].isHl:
                self.led[row].hours.append(Dhhl)
        elif type == "L":
            self.led[row].hours.append(0)
            # print('is tagged: ', self.led[row].available[col])
        else:
            self.led[row].hours.append(0)
            self.led[row].available[col] = True

    def calc_total_hours(self):
        for led in self.led:
            totHours = sum(led.hours)
            led.hours_tot = totHours

    """ Terminal outputs """

    def output(self):
        """ Prints to terminal """
        counter = 0
        print("\n")
        for i in self.led:
            # if i.isHl is False:
            print(
                f"{self.led[counter].name}:    {self.led[counter].total}      Total hours: {self.led[counter].hours_tot}")
            counter += 1

    def output_aval(self):
        """ Prints to terminal (availability) """
        counter = 0
        print("\n")
        for i in self.led:
            # if i.isHl is False:
            print(
                f"{self.led[counter].name}:    {self.led[counter].available}      Avaliable: {sum(self.led[counter].available)}")
            counter += 1

    def output_hours(self):
        """ Prints to terminal (Hours for the day) """
        counter = 0
        print("\n")
        for i in self.led:
            # if i.isHl is False:
            print(
                f"{self.led[counter].name}:    {self.led[counter].hours}      Avaliable: {self.led[counter].hours_tot}")
            counter += 1

    def save(self):
        """ Unused """
        self.wb.save(cfile)
